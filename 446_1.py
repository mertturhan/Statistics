import random
import numpy as np
import statistics as st
from matplotlib import pyplot as plt
import openpyxl as op
import datetime
import pandas as pd
import scipy.stats as stat
import math

#Questions 1-3

s=random.randint(1,1000)
xxx=random.seed(s)

n=random.randint(35,4000)

mu=np.random.normal(loc=2, scale=5, size=1)
sd=np.random.normal(loc=2, scale=5, size=1)


data=np.random.normal(loc=mu, scale=abs(sd), size=n)
agg=sum(data)


print(f"the average is = {agg/n}")
print(f"the variance is = {st.variance(data)}")
print(f"the median is = {st.median(data)}")

print(f"the random chosen mean and variance is respectively ={mu},{sd*sd}")

print(f"s={s}")
print(f"n={n}")

plt.hist(data, bins=20)
plt.show()

data2=np.random.normal(loc=mu, scale=abs(sd*2.5), size=n)

plt.hist(data2, bins=20)
plt.show()
print("QUESTION4")
path = "2022_HW1.xlsx"
wb = op.load_workbook(path)
m=wb['panel']

row = m.max_row
column = m.max_column

agg=0

for i in range(1, row + 1):
    try:
        agg+=float(m.cell(row = i, column = 2).value)
    except:
        continue

row_raw=0
for i in range(1, row + 1):
    if type(m.cell(row = i, column = 2).value)==float:
        row_raw+=1
    else:
        continue

meanr=(agg/row_raw)

print(f"The mean of return is={meanr}")

ret = [float(m.cell(row = i, column = 2).value) for i in range(1, row+1) if type(m.cell(row = i, column = 2).value)== float]

print(f"the median of return is = {st.median(ret)}")
print(f"the mode of return is = {st.mode(ret)}")
print(f"the max and min value of return is respectively = {max(ret)},{min(ret)}")

beta = [float(m.cell(row = i, column = 3).value) for i in range(1, row+1) if type(m.cell(row = i, column = 3).value)== float]

print(f"the mean of beta is = {st.mean(beta)}")
print(f"the median of beta is = {st.median(beta)}")
print(f"the mode of beta is = {st.mode(beta)}")
print(f"the max and min value of beta is respectively = {max(beta)},{min(beta)}")

beta_2006 = [float(m.cell(row = i, column = 3).value) for i in range(1, row+1) if type(m.cell(row = i, column = 3).value)== float and (m.cell(row = i, column = 4).value)==2006]
return_2006 = [float(m.cell(row = i, column = 2).value) for i in range(1, row+1) if type(m.cell(row = i, column = 2).value)== float and (m.cell(row = i, column = 4).value)==2006]

plt.hist(beta_2006, bins=40)
plt.show()
plt.hist(return_2006, bins=100)
plt.show()

#Question 5
h=wb['Houseprices']
row = h.max_row
column = h.max_column

vars=[]
yr=[]
for i in range(2, row+1):
        v=h.cell(row=i, column=1).value
        w=h.cell(row=i+1, column=1).value
        t=h.cell(row=i-1, column=2).value


        if type(w)!=datetime.datetime:
            yr.append(int(h.cell(row=i, column=2).value))
            print(f"Variance for Year {v.year} is = {int(st.variance(yr))}")
            vars.append(int(st.variance(yr)))
            yr.clear
            dataframe = pd.DataFrame({'date_of_year': np.array([datetime.datetime(a + 1990,1,1) for a in range(1, 18)]), 'variances': vars})
            plt.plot(dataframe.date_of_year, dataframe.variances)
            plt.title('Variances by years')
            plt.show()
        elif v.year == w.year:
            a=int(h.cell(row=i, column=2).value)
            yr.append(a)



        else:
            b=int(h.cell(row=i, column=2).value)
            yr.append(b)
            print(f"Variance for Year {v.year} is = {int(st.variance(yr))}")
            vars.append(int(st.variance(yr)))
            yr.clear()

#Question 6
h=wb['inflexch']
print(wb.sheetnames)
roww = h.max_row
columnn = h.max_column

def inflation(value1,value2):
    return np.log(value1)-np.log(value2)

for i in range(2,90):
    v=h.cell(row=i, column=2).value
    w=h.cell(row=i+12, column=2).value
    print(f"the yearly inflation for {i+10}th month is {inflation(float(w), float(v))}")

USD=[]
USD_lag1=[]
USD_lag2=[]
USD_lag3=[]
USD_lag4=[]
USD_lag5=[]
USD_lag6=[]
USD_lag7=[]
USD_lag8=[]
USD_lag9=[]
USD_lag10=[]
USD_lag11=[]
USD_lag12=[]
inf=[]
for i in range (2,102):
    t=h.cell(row=i, column=4).value
    z=h.cell(row=i, column=2).value
    bir=h.cell(row=i-1, column=4).value
    USD.append(t)
    if i>=3:
        USD_lag1.append(bir)
    if i>=4:
        iki = h.cell(row=i - 2, column=4).value
        USD_lag2.append(iki)
    if i>=5:
        uc = h.cell(row=i - 3, column=4).value
        USD_lag3.append(uc)
    if i>=6:
        dort = h.cell(row=i - 4, column=4).value
        USD_lag4.append(dort)
    if i>=7:
        bes = h.cell(row=i - 5, column=4).value
        USD_lag5.append(bes)
    if i>=8:
        alti = h.cell(row=i - 6, column=4).value
        USD_lag6.append(alti)
    if i>=9:
        yedi = h.cell(row=i - 7, column=4).value
        USD_lag7.append(yedi)
    if i>=10:
        sekiz = h.cell(row=i - 8, column=4).value
        USD_lag8.append(sekiz)
    if i>=11:
        dokuz = h.cell(row=i - 9, column=4).value
        USD_lag9.append(dokuz)
    if i>=12:
        on = h.cell(row=i - 10, column=4).value
        USD_lag10.append(on)
    if i>=13:
        onbir = h.cell(row=i - 11, column=4).value
        USD_lag11.append(onbir)
    if i>=14:
        oniki = h.cell(row=i - 12, column=4).value
        USD_lag12.append(oniki)

    inf.append(z)



USD_array=np.array(USD)
USD_lag1_array=np.array(USD_lag1)
USD_lag2_array=np.array(USD_lag2)
USD_lag3_array=np.array(USD_lag3)
USD_lag4_array=np.array(USD_lag4)
USD_lag5_array=np.array(USD_lag5)
USD_lag6_array=np.array(USD_lag6)
USD_lag7_array=np.array(USD_lag7)
USD_lag8_array=np.array(USD_lag8)
USD_lag9_array=np.array(USD_lag9)
USD_lag10_array=np.array(USD_lag10)
USD_lag11_array=np.array(USD_lag11)
USD_lag12_array=np.array(USD_lag12)
inf_array=np.array(inf)

corr, _ = stat.pearsonr(USD_array, inf_array)
print('Pearsons correlation for 0 lag: %.3f' % corr)

corr1, _ = stat.pearsonr(USD_lag1_array, inf_array[:-1])
print('Pearsons correlation for a month lag: %.3f' % corr1)

corr2, _ = stat.pearsonr(USD_lag2_array, inf_array[:-2])
print('Pearsons correlation for 2 months lag: %.3f' % corr2)

corr3, _ = stat.pearsonr(USD_lag3_array, inf_array[:-3])
print('Pearsons correlation for 3 months lag: %.3f' % corr3)

corr4, _ = stat.pearsonr(USD_lag4_array, inf_array[:-4])
print('Pearsons correlation for 4 months lag: %.3f' % corr2)

corr5, _ = stat.pearsonr(USD_lag5_array, inf_array[:-5])
print('Pearsons correlation for 5 months lag: %.3f' % corr5)

corr6, _ = stat.pearsonr(USD_lag6_array, inf_array[:-6])
print('Pearsons correlation for 6 months lag: %.3f' % corr6)

corr7, _ = stat.pearsonr(USD_lag7_array, inf_array[:-7])
print('Pearsons correlation for 7 months lag: %.3f' % corr7)

corr8, _ = stat.pearsonr(USD_lag8_array, inf_array[:-8])
print('Pearsons correlation for 8 months lag: %.3f' % corr8)

corr9, _ = stat.pearsonr(USD_lag9_array, inf_array[:-9])
print('Pearsons correlation for 9 months lag: %.3f' % corr5)

corr10, _ = stat.pearsonr(USD_lag10_array, inf_array[:-10])
print('Pearsons correlation for 10 months lag: %.3f' % corr6)

corr11, _ = stat.pearsonr(USD_lag11_array, inf_array[:-11])
print('Pearsons correlation for 11 months lag: %.3f' % corr11)

corr12, _ = stat.pearsonr(USD_lag12_array, inf_array[:-12])
print('Pearsons correlation for 12 months lag: %.3f' % corr12)

#Question 7
path = "pwt100.xlsx"
wb = op.load_workbook(path)
h=wb['Data']

roww = h.max_row
columnn = h.max_column
netherlands_gdp=[]
senegal_gdp=[]
for i in range(8872,8892):
    v=h.cell(row=i, column=5).value
    w=h.cell(row=i, column=7).value
    y=h.cell(row=i, column=4).value
    c=h.cell(row=i, column=3).value
    gdppc=v/w
    netherlands_gdp.append(gdppc)


for i in range(10202,10222):
    v=h.cell(row=i, column=5).value
    w=h.cell(row=i, column=7).value
    y=h.cell(row=i, column=4).value
    c=h.cell(row=i, column=3).value
    gdppc=v/w
    senegal_gdp.append(gdppc)


def rateofchange(year1, year2):
    return ((year2-year1)/(year1))

netherlands_gdp_change=[]

for i in netherlands_gdp:
    netherlands_gdp_change.append(rateofchange(i, i+1))

senegal_gdp_change=[]
for i in senegal_gdp:
    senegal_gdp_change.append(rateofchange(i, i+1))

neth = pd.DataFrame({'date_of_year': np.array([datetime.datetime(a + 1999, 1, 1) for a in range(1, 21)]), 'change': netherlands_gdp_change})
sen  = pd.DataFrame({'date_of_year': np.array([datetime.datetime(a + 1999, 1, 1) for a in range(1, 21)]), 'change': senegal_gdp_change})
plt.plot(neth.date_of_year, neth.change)
plt.title('Netherlands')
plt.xlabel('Year')
plt.ylabel('GDP Growth Rate')
plt.show()




plt.plot(neth.date_of_year, neth.change, label='Netherlands')
plt.plot(sen.date_of_year, sen.change, label='Senegal')
plt.title('Netherlands-Senegal GDP Growth rate comparison')
plt.show()

plt.plot(sen.date_of_year, sen.change)
plt.title('Senegal')
plt.xlabel('Year')
plt.ylabel('GDP Growth Rate')
plt.show()
plt.show()

#Question 8
def birthdayprob(n, m):
    formula = math.comb(n,m) * math.factorial(365) / (365**n * math.factorial(364-(n-m)))
    return formula

def atleast(n, m):
    formula = math.comb(n,m) * math.factorial(365) / (365**n * math.factorial(364-(n-m)))
    sum=0
    for i in range (m, n+1):
        sum+=birthdayprob(n, i)
    return sum

print(atleast(10,2))

fig = plt.figure()

# syntax for 3-D projection
ax = plt.axes(projection='3d')
# defining all 3 axes
for i in range (0,100):
     for j in range(0, i):
            X=i
            Y=j
            z= atleast(X,Y)
            if (z <= 1):
                ax.scatter(X, Y, z, 'green')


# plotting

ax.set_title('3D scatter plot')
plt.show()




#Question 9
n1=random.randint(100,200)
n2=random.randint(1000,3000)
n3=random.randint(4,8)
n4=random.randint(100,120)


violate_p = n1/n2

violaten_p=stat.binom.pmf(n3, n4 ,violate_p)

print(violaten_p)
#Question 10
print("Question 10")
n=17

x=np.arange(0,n+1)

binomial_25=stat.binom.pmf(x, n, 0.25)

plt.plot(x, binomial_25)
plt.show()

binomial_50=stat.binom.pmf(x, n, 0.5)

plt.plot(x, binomial_50)
plt.show()

binomial_75=stat.binom.pmf(x, n, 0.75)

plt.plot(x, binomial_75)
plt.show()

#Question 11
n=14

x=np.arange(0,n+1)

var = []



prob = np.random.uniform(low=0, high=1, size=10000)


for i in prob:

    binomial_pmf=stat.binom.pmf(x, n, i)
    damn=(i,st.variance(binomial_pmf))
    var.append(damn)



testList2 = [(e1, e2) for e1, e2 in var]


plt.scatter(*zip(*testList2),0,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,.2)
plt.title('Variance of a binomial distribution n=14')
plt.xlabel('Probability')
plt.ylabel('Variance')
plt.show()
