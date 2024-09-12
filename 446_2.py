import openpyxl as op
import scipy.stats as stat
import random
import numpy as np
from matplotlib import pyplot as plt
import statistics as st

path = "pwt100.xlsx"
wb = op.load_workbook(path)
h=wb['Data']
roww = h.max_row
columnn = h.max_column

inflation = []
exchange = []
lag_1 = []
lag_2 = []
lag_3 = []

for i in range (387, 403):
    ex=h.cell(row=i, column=29).value
    inf=h.cell(row=i, column=30).value
    inflation.append(inf)
    exchange.append(ex)
    bir = h.cell(row=i - 1, column=30).value
    if i>=388:
        lag_1.append(bir)
    if i>=389:
        iki = h.cell(row=i - 2, column=4).value
        lag_2.append(iki)
    if i>=390:
        uc = h.cell(row=i - 3, column=4).value
        lag_3.append(uc)
inflation_array=np.array(inflation)
exchange_array=np.array(exchange)
lag_1_array=np.array(lag_1)
lag_2_array=np.array(lag_2)
lag_3_array=np.array(lag_3)

corr, _ = stat.pearsonr(inflation_array, exchange_array)
print('Pearsons correlation between inflation and exchange rate: %.3f' % corr)

print("The correlation between exchange rate is high enough to say that they are correlated.")

corr1, _ = stat.pearsonr(lag_1_array, exchange_array[:-1])
print('Pearsons correlation between inflation and a month lag for exchange rate: %.3f' % corr1)

corr2, _ = stat.pearsonr(lag_2_array, exchange_array[:-2])
print('Pearsons correlation between inflation and 2 months lag for exchange rate: %.3f' % corr2)

corr3, _ = stat.pearsonr(lag_3_array, exchange_array[:-3])
print('Pearsons correlation between inflation and 3 months lag for exchange rate: %.3f' % corr3)

print("However, if we calculate the correlation between the exchange rate of 3 months ago from the given date and that date's inflation,we can deduce that they are HIGHLY correlated. The reason for this might be that Argentina is a highly import-depended country, that is Argentininan people have too much goods produced outside of Argentina in their consumer basket. Other reason for this correlation could be simply that Argentina might had increased their money supply by vast amounts, which naturally results in inflation if not followed by the same increase in aggregate demand by Argentinian people. The increase in money supply also makes the peso less dearer, which results in higher exchange rates")


aus_inf= []
nz_inf= []
for i in range (602, 613):
    inf = h.cell(row=i, column=30).value
    aus_inf.append(inf)

for i in range (9072,9083):
    inf = h.cell(row=i, column=30).value
    nz_inf.append(inf)

nz_inf_array=np.array(nz_inf)
aus_inf_array=np.array(aus_inf)


corrr, _ = stat.pearsonr(aus_inf_array, nz_inf_array)
print('Pearsons correlation between the inflation rate of Australia and New Zealand: %.3f' % corrr)

print("We deduce from this number that the inflation rate of inflation in Australia and New Zealand is HIGHLY correlated. The reason for this might stem from their geographical closeness. What this means is that they might be consuming similar basket of goods because of the shared environment and they might had followed similar fiscal and monetary policies")

######### 2nd question

theta = random.uniform(0,1)

print('%.3f'%theta)

n=100

x=np.arange(0,n+1)

binomial_cdf = stat.binom.cdf(x, n, theta)



plt.plot(x, binomial_cdf, color='blue')
plt.title(f"Binomial Cumulative Distribution(n={n}, p={'%.3f'%theta})")
plt.show()


zlist= []

for i in x:
    z=(i-n*theta)/((n*theta*(1-theta))**1/2)
    zlist.append(z)

mu= n*theta
sigma = (n*theta*(1-theta))**1/2



normal_cdf = stat.norm.cdf(mu, sigma)

normal_cdf = []

for i in zlist:
    normal_cdf.append(stat.norm(mu, sigma).cdf(i))


plt.plot(x, normal_cdf, color='blue')
plt.show()

delta=[]

delta=[abs(a - b) for b, a in zip(normal_cdf, binomial_cdf)]

plt.plot(x, delta)
plt.show()


print("The way we try to prove the statement is definitely false since we get the value 1 in the both distribution as we get closer to the end of array n. Hence the third process gives us 0 no matter the magnitude of n.  ")

##### 3rd question

n=25000
x=np.arange(n+1)
uni_pdf=stat.uniform.pdf(x,0,n)

plt.plot(x, uni_pdf)
plt.show()

a=random.randint(10000,n)

uni_pdf_l = np.array(uni_pdf)
list_1 = uni_pdf_l.tolist()
s1=random.sample(list_1,a)

b=random.randint(5,a-15)
s2=random.sample(s1,b)

print(st.mean(s2))

s3 = s2
print(st.mean(s2))
for i in range (b, b+12):
    a=random.sample(s1,1)
    s3.append(a[0])
    print(st.mean(s3))
    print(stat.kurtosis(s3))